from fastapi import FastAPI, HTTPException, Depends
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from sqlalchemy.orm import Session
from typing import Optional
import os
import logging
from dotenv import load_dotenv
load_dotenv(override=True)

logger = logging.getLogger(__name__)

from app.models import (
    CheckRequest, CheckResponse, BlockResponse, RedactedType,
    SessionListResponse, SessionDetailResponse, ChatSessionInfo, ChatMessageInfo,
    TierConfigUpdate, TierConfigResponse, StatsResponse, StatCount,
    OrgModelConfig, ApiKeyCreate
)
from app.pipeline import pipeline
from app.pipeline import regex_stage
from app.config import get_block_warning, TIER_BLOCK, TIER_REDACT, TIER_AUDIT
from app import llm_client
from app import tokenization
from app.models_db import init_db, get_db, Session as ChatSession, Message as ChatMessage, User, StatLog, CustomLabel, Organization
import json
from collections import Counter
from sqlalchemy import func
from fastapi.responses import StreamingResponse
from google.genai import types
import google.genai as genai
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import io
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from fastapi import Request
import asyncio
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import jwt
from jwt import PyJWKClient

security_bearer = HTTPBearer(auto_error=False)


CLERK_JWKS_URL = os.getenv("CLERK_JWKS_URL", "https://rapid-hyena-22.clerk.accounts.dev/.well-known/jwks.json")
try:
    jwks_client = PyJWKClient(CLERK_JWKS_URL, cache_keys=True, lifespan=3600)
except Exception as e:
    logger.critical("JWKS client failed to initialise — all JWT auth will fail: %s", e)
    jwks_client = None

_GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip('"').strip("'")
_gemini_client = genai.Client(api_key=_GEMINI_API_KEY) if _GEMINI_API_KEY else None


_NOAUTH_DEV = os.getenv("NOAUTH_DEV", "").lower() in ("1", "true", "yes")
if _NOAUTH_DEV and os.getenv("ENVIRONMENT", "").lower() == "production":
    raise RuntimeError("NOAUTH_DEV must not be enabled in production")


def _validate_regex_safe(pattern: str, timeout: float = 0.5) -> None:
    """Compile and probe a regex for catastrophic backtracking (ReDoS) before saving."""
    import re as _re, threading
    try:
        compiled = _re.compile(pattern)
    except _re.error as e:
        raise HTTPException(status_code=400, detail=f"Invalid regex pattern: {e}")
    done = []
    def _probe():
        try:
            compiled.search("a" * 100 + "!")
            done.append(True)
        except Exception:
            done.append(False)
    t = threading.Thread(target=_probe, daemon=True)
    t.start()
    t.join(timeout)
    if not done:
        raise HTTPException(status_code=400, detail="Regex pattern timed out — potential ReDoS risk. Simplify the pattern.")

def verify_credentials(request: Request, bearer_creds: HTTPAuthorizationCredentials = Depends(security_bearer), db: Session = Depends(get_db)):
    from app.models_db import Organization, User

    if _NOAUTH_DEV:
        user = db.query(User).filter(User.email == "dev@local").first()
        if not user:
            user = User(email="dev@local", role="admin", password_hash="dev", is_active=True, rate_limit_per_day=100)
            db.add(user)
            db.commit()
            db.refresh(user)
        user.is_base_user = True
        request.state.is_base_user = True
        request.state.current_user = user
        request.state.clerk_org_id = None
        request.state.clerk_user_id = "dev@local"
        return True

    if bearer_creds and bearer_creds.credentials.startswith("adpsh_"):
        import hashlib
        from datetime import datetime, timezone
        from app.models_db import ApiKey
        key_hash = hashlib.sha256(bearer_creds.credentials.encode()).hexdigest()
        api_key = db.query(ApiKey).filter(ApiKey.key_hash == key_hash, ApiKey.is_active == True).first()
        if not api_key:
            raise HTTPException(status_code=401, detail="Invalid API key")
        if api_key.expires_at and api_key.expires_at < datetime.now(timezone.utc):
            raise HTTPException(status_code=401, detail="API key expired")

        key_user = db.query(User).filter(User.id == api_key.user_id).first()
        if not key_user or getattr(key_user, "is_blocked", False):
            raise HTTPException(status_code=401, detail="API key owner is inactive")

        api_key.last_used_at = datetime.now(timezone.utc)
        api_key.last_used_ip = get_remote_address(request)
        db.commit()

        key_user.is_base_user = (key_user.org_id is None)
        request.state.is_base_user = key_user.is_base_user
        request.state.current_user = key_user
        request.state.via_api_key = True
        request.state.api_key_scopes = api_key.scopes or []
        request.state.api_key_id = api_key.id
        request.state.clerk_org_id = None
        request.state.clerk_user_id = str(key_user.id)

        if request.url.path.startswith("/api/v1/admin"):
            raise HTTPException(status_code=403, detail="API keys cannot access admin endpoints")

        limit = api_key.rate_limit_per_min or 60
        remaining = enforce_api_rate_limit(str(api_key.id), limit)
        request.state.rl_limit = limit
        request.state.rl_remaining = remaining
        request.state.rl_reset = int(_time.time()) + 60
        return True

    try:
        user = None

        global jwks_client
        if not jwks_client:
            try:
                jwks_client = PyJWKClient(CLERK_JWKS_URL, cache_keys=True, lifespan=3600)
            except Exception as e:
                print("Failed to initialize JWKS client:", e)


        if not user and bearer_creds and jwks_client:
            token = bearer_creds.credentials
            try:
                signing_key = jwks_client.get_signing_key_from_jwt(token)
                data = jwt.decode(
                    token,
                    signing_key.key,
                    algorithms=["RS256"],
                    options={"verify_aud": False},
                    leeway=30
                )
                import uuid
                clerk_id = data.get("sub")
                clerk_org_id = data.get("org_id")
                
                if clerk_org_id:
                    CLERK_NAMESPACE = uuid.uuid5(uuid.NAMESPACE_DNS, 'clerk.adopshun.com')  # deterministic: same Clerk org always maps to the same UUID
                    org_uuid = uuid.uuid5(CLERK_NAMESPACE, clerk_org_id)
                    org = db.query(Organization).filter(Organization.id == org_uuid).first()
                    if not org:
                        org_slug = data.get("org_slug", "Unknown Organization")
                        org = Organization(
                            id=org_uuid,
                            name=org_slug.replace("-", " ").title() if org_slug else "New Organization",
                            is_active=True
                        )
                        db.add(org)
                else:
                    org = None

                clerk_email = data.get("email") or data.get("email_address") or ""  # field name depends on Clerk JWT template config

                user = db.query(User).filter(User.email == clerk_id).first()
                if not user:
                    user = User(
                        org_id=org.id if org else None,
                        email=clerk_id,
                        employee_id=clerk_email or None,
                        role="admin" if data.get("org_role") == "org:admin" or not clerk_org_id else "employee",
                        password_hash="clerk_managed",
                        is_active=True,
                        rate_limit_per_day=15 if not clerk_org_id else None
                    )
                    db.add(user)
                    db.commit()
                    db.refresh(user)
                else:
                    if clerk_org_id and org and user.org_id != org.id:
                        user.org_id = org.id
                        user.role = "admin" if data.get("org_role") == "org:admin" else "employee"
                        user.rate_limit_per_day = None
                    if clerk_email and not user.employee_id:
                        user.employee_id = clerk_email
                    db.commit()
                user.is_base_user = (user.org_id is None)
                request.state.clerk_org_id = clerk_org_id
                request.state.clerk_user_id = clerk_id
                if getattr(user, 'is_blocked', False):
                    raise HTTPException(status_code=403, detail="Account suspended")
            except HTTPException:
                raise
            except Exception as e:
                logger.error("Clerk JWT Decode Error: %s", e, exc_info=True)
                raise HTTPException(status_code=401, detail="Authentication failed")
                
        if not user:
            print("Auth Failed: User not found or bearer token missing.")
            if not bearer_creds:
                print("- No Bearer token provided in headers")
            raise HTTPException(
                status_code=401,
                detail="Not authenticated",
                headers={"WWW-Authenticate": "Bearer"},
            )
            
        request.state.is_base_user = getattr(user, 'is_base_user', True)
            
        if request.url.path.startswith("/api/v1/admin") and user.role not in ["admin", "super_admin"]:
            raise HTTPException(
                status_code=403,
                detail="Admin privileges required"
            )
        request.state.current_user = user
        return True
    finally:
        pass


def require_scope(request: Request, scope: str):
    """Enforce API-key scopes. No-op for Clerk-authenticated (browser) requests."""
    if getattr(request.state, "via_api_key", False):
        if scope not in (getattr(request.state, "api_key_scopes", None) or []):
            raise HTTPException(status_code=403, detail=f"API key missing required scope: {scope}")


import time as _time
import threading as _threading
_rate_buckets: dict = {}  # per-process; resets on restart — Redis upgrade path for multi-worker
_rate_lock = _threading.Lock()

def enforce_api_rate_limit(key_id: str, limit_per_min: int) -> int:
    """Records a hit; returns remaining requests in the window. Raises 429 if over."""
    now = _time.time()
    cutoff = now - 60
    with _rate_lock:
        bucket = [t for t in _rate_buckets.get(key_id, []) if t > cutoff]
        if len(bucket) >= limit_per_min:
            _rate_buckets[key_id] = bucket
            raise HTTPException(
                status_code=429,
                detail=f"Rate limit exceeded ({limit_per_min}/min for this API key)",
                headers={"Retry-After": "60", "X-RateLimit-Limit": str(limit_per_min), "X-RateLimit-Remaining": "0"},
            )
        bucket.append(now)
        _rate_buckets[key_id] = bucket
        return max(0, limit_per_min - len(bucket))


_idem_cache: dict = {}  # per-process 24 h idempotency cache; Redis upgrade path for multi-worker
_idem_lock = _threading.Lock()
_IDEM_TTL = 24 * 3600

def idem_get(cache_key):
    with _idem_lock:
        v = _idem_cache.get(cache_key)
        if v and v[0] > _time.time():
            return v[1]
        if v:
            _idem_cache.pop(cache_key, None)
    return None

def idem_set(cache_key, response: dict):
    with _idem_lock:
        _idem_cache[cache_key] = (_time.time() + _IDEM_TTL, response)


def fire_webhook(org, event: dict):
    """
    Fire-and-forget HMAC-signed webhook for security events (e.g. BLOCK).
    No-op if the org has no webhook configured. Never raises into the request.
    """
    url = getattr(org, "webhook_url", None) if org else None
    if not url:
        return
    secret = (getattr(org, "webhook_secret", None) or "").encode()

    def _send():
        try:
            import hmac, hashlib, requests
            body = json.dumps(event).encode()
            sig = hmac.new(secret, body, hashlib.sha256).hexdigest() if secret else ""
            requests.post(
                url,
                data=body,
                headers={"Content-Type": "application/json", "X-Adopshun-Signature": f"sha256={sig}"},
                timeout=5,
            )
        except Exception as e:
            logger.warning("Webhook delivery failed: %s", e)

    _threading.Thread(target=_send, daemon=True).start()


def purge_old_data(db, org_id=None) -> int:
    """
    Delete stat logs, sessions and messages older than each org's retention_days.
    Scoped to one org when org_id is given. Skips orgs with no retention set.
    ponytail: invoked on startup and on demand; the production path is a daily cron.
    """
    from datetime import datetime, timezone, timedelta
    orgs = db.query(Organization).filter(Organization.id == org_id).all() if org_id else db.query(Organization).all()
    deleted = 0
    for org in orgs:
        days = org.retention_days or 0
        if days <= 0:
            continue
        cutoff = datetime.now(timezone.utc) - timedelta(days=days)
        deleted += db.query(StatLog).filter(
            StatLog.org_id == org.id, StatLog.created_at < cutoff
        ).delete(synchronize_session=False)
        old = db.query(ChatSession).filter(
            ChatSession.org_id == org.id, ChatSession.created_at < cutoff
        ).all()
        for s in old:
            db.query(ChatMessage).filter(ChatMessage.session_id == s.id).delete(synchronize_session=False)
        db.query(ChatSession).filter(
            ChatSession.org_id == org.id, ChatSession.created_at < cutoff
        ).delete(synchronize_session=False)
    db.commit()
    return deleted


def validate_external_url(url: str):
    """
    Block SSRF on admin-supplied LLM endpoints: only http(s), and the host must
    not resolve to a private/loopback/link-local/reserved address (e.g. the cloud
    metadata endpoint 169.254.169.254 or internal services).
    ponytail: validated at set-time only — DNS rebinding (TOCTOU) is not covered;
    upgrade path is to re-resolve and pin the IP at request time.
    """
    import socket, ipaddress
    from urllib.parse import urlparse

    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        raise HTTPException(status_code=400, detail="Endpoint URL must use http or https")
    host = parsed.hostname
    if not host:
        raise HTTPException(status_code=400, detail="Invalid endpoint URL")
    try:
        infos = socket.getaddrinfo(host, None)
    except socket.gaierror:
        raise HTTPException(status_code=400, detail="Endpoint host does not resolve")
    for info in infos:
        ip = ipaddress.ip_address(info[4][0])
        if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved or ip.is_multicast or ip.is_unspecified:
            raise HTTPException(status_code=400, detail="Endpoint resolves to a disallowed internal address")


limiter = Limiter(key_func=get_remote_address)
app = FastAPI(
    title="PII Detection API",
    description="API for detecting and anonymizing PII data",
)

from fastapi.middleware.cors import CORSMiddleware

_origins = os.getenv("ALLOWED_ORIGINS", "http://localhost:3000,http://127.0.0.1:3000,https://pii-api-mu.vercel.app").split(",")
if "https://chat.adopshun.com" not in _origins:
    _origins.append("https://chat.adopshun.com")

app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def rate_limit_headers(request: Request, call_next):
    response = await call_next(request)
    if hasattr(request.state, "rl_limit"):
        response.headers["X-RateLimit-Limit"] = str(request.state.rl_limit)
        response.headers["X-RateLimit-Remaining"] = str(request.state.rl_remaining)
        response.headers["X-RateLimit-Reset"] = str(request.state.rl_reset)
    return response

init_db()


@app.on_event("startup")
def _startup_retention_purge():
    """Best-effort retention enforcement on boot. A daily cron is the real mechanism."""
    try:
        db = next(get_db())
        n = purge_old_data(db)
        if n:
            print(f"Retention purge removed {n} stat-log rows.")
        db.close()
    except Exception as e:
        print("Retention purge skipped:", e)


def get_tier_config_helper(db, user):
    is_base = getattr(user, "is_base_user", True)
    if user:
        if is_base:
            custom_labels = db.query(CustomLabel).filter(CustomLabel.user_id == user.id).all()
        else:
            custom_labels = db.query(CustomLabel).filter(CustomLabel.org_id == user.org_id).all()
    else:
        custom_labels = []
        
    if user and user.tier_block:
        tier_config = {"block": set(user.tier_block), "redact": set(user.tier_redact), "audit": set(user.tier_audit)}
    else:
        if user and not is_base:
            org = db.query(Organization).filter(Organization.id == user.org_id).first()
            if org and org.default_tier_block:
                tier_config = {
                    "block": set(org.default_tier_block),
                    "redact": set(org.default_tier_redact),
                    "audit": set(org.default_tier_audit)
                }
                already_tiered = tier_config["block"] | tier_config["redact"] | tier_config["audit"]
                for cl in custom_labels:
                    if cl.name in already_tiered:
                        continue
                    if cl.tier == "tier_block":
                        tier_config["block"].add(cl.name)
                    elif cl.tier == "tier_redact":
                        tier_config["redact"].add(cl.name)
                    elif cl.tier == "tier_audit":
                        tier_config["audit"].add(cl.name)
                return tier_config, custom_labels
        tier_config = {"block": set(TIER_BLOCK), "redact": set(TIER_REDACT), "audit": set(TIER_AUDIT)}

    already_tiered = tier_config["block"] | tier_config["redact"] | tier_config["audit"]
    for cl in custom_labels:
        if cl.name in already_tiered:
            continue
        if cl.tier == "tier_block":
            tier_config["block"].add(cl.name)
        elif cl.tier == "tier_redact":
            tier_config["redact"].add(cl.name)
        elif cl.tier == "tier_audit":
            tier_config["audit"].add(cl.name)
    return tier_config, custom_labels


app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    import traceback
    err = traceback.format_exc()
    print("GLOBAL EXCEPTION:", err)
    response = JSONResponse(status_code=500, content={"detail": "Internal Server Error"})
    origin = request.headers.get("origin")
    allowed_origins = set(os.getenv("ALLOWED_ORIGINS", "http://localhost:3000,http://127.0.0.1:3000").split(","))
    allowed_origins.add("https://chat.adopshun.com")
    if origin and origin in allowed_origins:
        response.headers["Access-Control-Allow-Origin"] = origin
    response.headers["Access-Control-Allow-Credentials"] = "true"
    response.headers["Access-Control-Allow-Methods"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "*"
    return response

@app.get("/")
def read_root():
    return {"status": "ok", "message": "PII Detection API is running"}

@app.post("/api/v1/check")
@limiter.limit("20/minute")
async def check_message(request: Request, body: CheckRequest, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    require_scope(request, "check")
    user = request.state.current_user

    idem_key = None
    if getattr(request.state, "via_api_key", False):
        hdr = request.headers.get("Idempotency-Key")
        if hdr:
            idem_key = (str(getattr(request.state, "api_key_id", "")), hdr)
            cached = idem_get(idem_key)
            if cached is not None:
                return JSONResponse(cached)

    tier_config, custom_labels = get_tier_config_helper(db, user)
    processed_text, detections, action = await asyncio.to_thread(pipeline.run, body.message, body.allowed_pii, body.ignored_values, tier_config, custom_labels)

    if getattr(request.state, "via_api_key", False):
        block_set = tier_config["block"] if tier_config else TIER_BLOCK
        db.add(StatLog(  # store types only — never raw PII values on the API path
            user_id=user.id,
            org_id=user.org_id,
            session_id=None,
            api_key_id=getattr(request.state, "api_key_id", None),
            action=action,
            detected_types=json.dumps([d.type for d in detections]),
            flagged_sequences=json.dumps([]),
            original_message=None,
        ))
        db.commit()
        tokenized, vault = tokenization.build_tokens(body.message, detections, block_set)  # vault returned to caller, never stored
        resp = {
            "action": action,
            "was_redacted": action == "REDACT",
            "message": processed_text,
            "tokenized": tokenized,
            "vault": vault,
            "detections": [
                {
                    "type": d.type,
                    "subtype": d.subtype,
                    "confidence": d.confidence,
                    "start": d.start,
                    "end": d.end,
                    "value": body.message[d.start:d.end],
                }
                for d in detections
            ],
        }
        if idem_key is not None:
            idem_set(idem_key, resp)
        if action == "BLOCK" and user.org_id:
            org = db.query(Organization).filter(Organization.id == user.org_id).first()
            fire_webhook(org, {"event": "pii_blocked", "source": "api", "detected_types": [d.type for d in detections], "user_id": str(user.id)})
        return JSONResponse(resp)

    import uuid
    try:
        real_sess_id = uuid.UUID(body.session_id)
    except:
        real_sess_id = None

    detected_types_list = [d.type for d in detections]
    flagged_sequences_list = [body.message[d.start:d.end] for d in detections]
    stat_log = StatLog(
        user_id=user.id,
        org_id=user.org_id,
        session_id=real_sess_id,
        action=action,
        detected_types=json.dumps(detected_types_list),
        flagged_sequences=json.dumps(flagged_sequences_list),
        original_message=body.message if action in ["BLOCK", "REDACT"] else None
    )
    db.add(stat_log)
    db.commit()
    
    redacted_types_dicts = [
        {
            "type": d.type, 
            "subtype": d.subtype, 
            "confidence": d.confidence,
            "value": body.message[d.start:d.end]
        }
        for d in detections
    ]
    
    if action == "BLOCK":
        block_set = tier_config["block"] if tier_config else TIER_BLOCK
        blocked_types = [d.type for d in detections if d.type in block_set]
        primary_type = blocked_types[0] if blocked_types else "code"

        if real_sess_id:
            db_sess = db.query(ChatSession).filter(ChatSession.id == real_sess_id).first()
            if not db_sess:
                _title = body.message[:30] + "..." if len(body.message) > 30 else body.message
                db_sess = ChatSession(id=real_sess_id, user_id=user.id, org_id=user.org_id, title=_title)
                db.add(db_sess)
            blocked_msg = ChatMessage(session_id=real_sess_id, role="blocked", content=body.message)
            db.add(blocked_msg)
            db.commit()
        
        blocked_redacted_types = [rt for rt in redacted_types_dicts if rt["type"] in block_set]

        if user.org_id:
            _org = db.query(Organization).filter(Organization.id == user.org_id).first()
            fire_webhook(_org, {"event": "pii_blocked", "source": "chat", "detected_types": blocked_types, "user_id": str(user.id)})

        raise HTTPException(
            status_code=400,
            detail=BlockResponse(
                action="BLOCK",
                warning=get_block_warning(primary_type),
                blocked_types=[RedactedType(**rt) for rt in blocked_redacted_types]
            ).model_dump()
        )

    org = db.query(Organization).filter(Organization.id == user.org_id).first() if user.org_id else None
    db_session = None
    if real_sess_id:
        db_session = db.query(ChatSession).filter(ChatSession.id == real_sess_id).first()
        if not db_session:
            title = body.message[:30] + "..." if len(body.message) > 30 else body.message
            db_session = ChatSession(id=real_sess_id, user_id=user.id, org_id=user.org_id, title=title)
            db.add(db_session)
            db.commit()

        user_msg = ChatMessage(
            session_id=real_sess_id,
            role="user",
            content=processed_text,
            redacted_types=json.dumps(redacted_types_dicts)
        )
        db.add(user_msg)
        db.commit()

    resolved_model = llm_client.resolve_model(body.model, session=db_session, user=user, org=org)
    if db_session and db_session.model_used != resolved_model:
        db_session.model_used = resolved_model
        db.commit()

    async def generate():
        metadata = {
            "type": "metadata",
            "action": action,
            "message": processed_text,
            "redacted_types": redacted_types_dicts
        }
        yield f"data: {json.dumps(metadata)}\n\n"

        llm_reply = ""
        available = llm_client.get_available_models(user, org)
        if available and real_sess_id:
            history = db.query(ChatMessage).filter(ChatMessage.session_id == real_sess_id).order_by(ChatMessage.created_at).all()
            prior_history = history[:-1]   # exclude the message we just stored

            try:
                usage_out = {}

                async def _collect_stream():
                    parts = []
                    async for text in llm_client.stream_response(
                        resolved_model, prior_history, processed_text, org=org, usage_out=usage_out
                    ):
                        parts.append(text)
                    return parts

                try:
                    chunks = await asyncio.wait_for(_collect_stream(), timeout=45)
                    for text in chunks:
                        llm_reply += text
                        yield f"data: {json.dumps({'type': 'chunk', 'text': text})}\n\n"
                except asyncio.TimeoutError:
                    yield f"data: {json.dumps({'type': 'error', 'text': 'Request timed out. Please try again.'})}\n\n"
                    return

                model_msg = ChatMessage(session_id=real_sess_id, role="model", content=llm_reply)
                db.add(model_msg)
                if usage_out.get("total_tokens"):  # NULL when stream was interrupted — flagged in dashboard
                    stat_log.tokens = usage_out["total_tokens"]
                    db.add(stat_log)
                db.commit()
            except Exception as e:
                yield f"data: {json.dumps({'type': 'error', 'text': f'Error calling model: {str(e)}'})}\n\n"
        else:
            llm_reply = "I am the pseudo LLM. I received your message securely. (No LLM provider configured)"
            yield f"data: {json.dumps({'type': 'chunk', 'text': llm_reply})}\n\n"

        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


@app.get("/api/v1/models")
def list_models(request: Request, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    user = request.state.current_user
    org = db.query(Organization).filter(Organization.id == user.org_id).first() if user.org_id else None
    return {"models": llm_client.get_available_models(user, org)}


@app.get("/api/v1/admin/models/config")
def get_org_model_config(request: Request, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    """Current org model settings plus the full catalogue an admin can choose from."""
    user = request.state.current_user
    org = db.query(Organization).filter(Organization.id == user.org_id).first() if user.org_id else None

    catalogue = [
        {"id": mid, "display": m["display"], "tier": m["tier"], "provider": m["provider"],
         "configured": bool(os.getenv(m["env"]))}
        for mid, m in llm_client.SUPPORTED_MODELS.items()
    ]
    cfg = (getattr(org, "llm_config", None) or {}) if org else {}
    custom = {k: v for k, v in cfg.items() if k != "api_key"} if cfg else None  # never return stored api_key to client
    return {
        "catalogue": catalogue,
        "default_model": getattr(org, "default_model", None) if org else None,
        "allowed_models": getattr(org, "allowed_models", None) or [] if org else [],
        "custom_endpoint": custom,
        "custom_configured": llm_client._custom_configured(org),
        "webhook_url": getattr(org, "webhook_url", None) if org else None,
        "webhook_configured": bool(getattr(org, "webhook_secret", None)) if org else False,
    }


@app.put("/api/v1/admin/models/config")
def update_org_model_config(request: Request, body: OrgModelConfig, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    user = request.state.current_user
    if not user.org_id:
        raise HTTPException(status_code=400, detail="Model configuration is an organization feature")

    org = db.query(Organization).filter(Organization.id == user.org_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    org.allowed_models = body.allowed_models
    org.default_model = body.default_model

    if body.custom_endpoint:
        ce = body.custom_endpoint
        validate_external_url(ce.base_url)  # SSRF guard
        existing = org.llm_config or {}
        org.llm_config = {
            "base_url": ce.base_url,
            "api_key": ce.api_key if ce.api_key else existing.get("api_key"),  # blank on edit = keep existing
            "model_name": ce.model_name,
            "display_name": ce.display_name,
        }

    if body.webhook_url is not None:
        if body.webhook_url:
            validate_external_url(body.webhook_url)  # SSRF guard — same as LLM endpoint
        org.webhook_url = body.webhook_url or None
    if body.webhook_secret:
        org.webhook_secret = body.webhook_secret

    db.add(org)
    db.commit()
    return {"status": "updated"}


VALID_API_SCOPES = {"check", "read:stats"}


def _api_key_to_info(k) -> dict:
    return {
        "id": str(k.id),
        "name": k.name,
        "prefix": k.prefix,
        "scopes": k.scopes or [],
        "rate_limit_per_min": k.rate_limit_per_min or 60,
        "last_used_at": k.last_used_at,
        "last_used_ip": k.last_used_ip,
        "expires_at": k.expires_at,
        "is_active": k.is_active,
        "created_at": k.created_at,
    }


@app.post("/api/v1/admin/api-keys")
def create_api_key(request: Request, body: ApiKeyCreate, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    import secrets, hashlib
    from datetime import datetime, timezone, timedelta
    from app.models_db import ApiKey

    user = request.state.current_user

    scopes = [s for s in body.scopes if s in VALID_API_SCOPES] or ["check"]
    raw = "adpsh_" + secrets.token_urlsafe(32)
    key_hash = hashlib.sha256(raw.encode()).hexdigest()
    expires_at = None
    if body.expires_in_days and body.expires_in_days > 0:
        expires_at = datetime.now(timezone.utc) + timedelta(days=body.expires_in_days)

    api_key = ApiKey(
        user_id=user.id,
        org_id=user.org_id,
        name=body.name,
        prefix=raw[:12],
        key_hash=key_hash,
        scopes=scopes,
        rate_limit_per_min=body.rate_limit_per_min,
        expires_at=expires_at,
        is_active=True,
    )
    db.add(api_key)
    db.commit()
    db.refresh(api_key)

    info = _api_key_to_info(api_key)
    info["key"] = raw   # shown exactly once
    return info


@app.get("/api/v1/admin/api-keys")
def list_api_keys(request: Request, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    from app.models_db import ApiKey
    from datetime import datetime, timezone, timedelta
    user = request.state.current_user
    keys = db.query(ApiKey).filter(ApiKey.user_id == user.id).order_by(ApiKey.created_at.desc()).all()

    totals, recent = {}, {}
    if keys:
        key_ids = [k.id for k in keys]
        day_ago = datetime.now(timezone.utc) - timedelta(hours=24)
        totals = dict(
            db.query(StatLog.api_key_id, func.count(StatLog.id))
            .filter(StatLog.api_key_id.in_(key_ids))
            .group_by(StatLog.api_key_id).all()
        )
        recent = dict(
            db.query(StatLog.api_key_id, func.count(StatLog.id))
            .filter(StatLog.api_key_id.in_(key_ids), StatLog.created_at >= day_ago)
            .group_by(StatLog.api_key_id).all()
        )

    out = []
    for k in keys:
        info = _api_key_to_info(k)
        info["total_calls"] = totals.get(k.id, 0)
        info["calls_24h"] = recent.get(k.id, 0)
        out.append(info)
    return {"api_keys": out}


@app.delete("/api/v1/admin/api-keys/{key_id}")
def revoke_api_key(request: Request, key_id: str, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    import uuid as _uuid
    from app.models_db import ApiKey
    user = request.state.current_user
    try:
        kid = _uuid.UUID(key_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid key id")

    key = db.query(ApiKey).filter(ApiKey.id == kid).first()
    if not key:
        raise HTTPException(status_code=404, detail="API key not found")
    if key.user_id != user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    key.is_active = False
    db.commit()
    return {"status": "revoked"}

from typing import Union
from app.models import BatchCheckRequest, BatchCheckResponse, CheckResult, BlockResult

@app.post("/api/v1/preview", response_model=Union[CheckResult, BlockResult])
@limiter.limit("30/minute")
async def preview_message(request: Request, body: CheckRequest, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    require_scope(request, "check")
    user = request.state.current_user
    tier_config, custom_labels = get_tier_config_helper(db, user)
    processed_text, detections, action = await asyncio.to_thread(pipeline.run, body.message, body.allowed_pii, body.ignored_values, tier_config, custom_labels)

    import uuid
    try:
        real_sess_id = uuid.UUID(body.session_id)
    except:
        real_sess_id = None

    detected_types_list = [d.type for d in detections]
    flagged_sequences_list = [body.message[d.start:d.end] for d in detections]
    stat_log = StatLog(
        user_id=user.id,
        org_id=user.org_id,
        session_id=real_sess_id,
        action=action,
        detected_types=json.dumps(detected_types_list),
        flagged_sequences=json.dumps(flagged_sequences_list)
    )
    db.add(stat_log)
    db.commit()
        
    redacted_types = [
        RedactedType(type=d.type, subtype=d.subtype, confidence=d.confidence, value=body.message[d.start:d.end])
        for d in detections
    ]
    
    if action == "BLOCK":
        block_set = tier_config["block"] if tier_config else TIER_BLOCK
        blocked_types = [d.type for d in detections if d.type in block_set]
        primary_type = blocked_types[0] if blocked_types else "code"
        
        blocked_redacted_types = [rt for rt in redacted_types if rt.type in block_set]
        
        return BlockResult(
            action="BLOCK",
            warning=get_block_warning(primary_type),
            blocked_types=blocked_redacted_types
        )
    
    return CheckResult(
        action=action,
        was_redacted=(action == "REDACT"),
        message=processed_text,
        redacted_types=redacted_types
    )
@app.get("/api/v1/sessions", response_model=SessionListResponse)
def get_sessions(request: Request, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    user = request.state.current_user
    is_base = getattr(request.state, "is_base_user", True)
    if is_base:
        sessions = db.query(ChatSession).filter(ChatSession.user_id == user.id).order_by(ChatSession.created_at.desc()).all()
    else:
        sessions = db.query(ChatSession).filter(ChatSession.org_id == user.org_id).order_by(ChatSession.created_at.desc()).all()
    return SessionListResponse(
        sessions=[ChatSessionInfo(id=str(s.id), title=s.title, created_at=s.created_at) for s in sessions]
    )

@app.get("/api/v1/sessions/{session_id}", response_model=SessionDetailResponse)
def get_session(request: Request, session_id: str, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    user = request.state.current_user
    is_base = getattr(request.state, "is_base_user", True)
    query = db.query(ChatSession).filter(ChatSession.id == session_id)
    if is_base:
        query = query.filter(ChatSession.user_id == user.id)
    else:
        query = query.filter(ChatSession.org_id == user.org_id)
    session = query.first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    messages = db.query(ChatMessage).filter(ChatMessage.session_id == session_id).order_by(ChatMessage.created_at).all()
    

    msg_infos = []
    for m in messages:
        rt = []
        if m.redacted_types:
            try:
                rt = json.loads(m.redacted_types)
            except: pass
        msg_infos.append(ChatMessageInfo(role=m.role, content=m.content, created_at=m.created_at, redacted_types=rt))
        
    return SessionDetailResponse(
        id=str(session.id),
        title=session.title,
        messages=msg_infos,
        model_used=session.model_used
    )

@app.delete("/api/v1/sessions/{session_id}")
def delete_session(request: Request, session_id: str, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    user = request.state.current_user
    is_base = getattr(request.state, "is_base_user", True)
    query = db.query(ChatSession).filter(ChatSession.id == session_id)
    if is_base:
        query = query.filter(ChatSession.user_id == user.id)
    else:
        query = query.filter(ChatSession.org_id == user.org_id)
    session = query.first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    db.query(ChatMessage).filter(ChatMessage.session_id == session_id).delete()
    db.delete(session)
    db.commit()
    return {"status": "deleted"}

from pydantic import BaseModel as _BaseModel

class BlockedMessageSave(_BaseModel):
    session_id: str
    message: str
    user_id: str = "default_user"
    model_explanation: Optional[str] = None
    blocked_types: Optional[list] = None

@app.post("/api/v1/sessions/save-blocked")
@limiter.limit("20/minute")
def save_blocked_message(request: Request, body: BlockedMessageSave, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    user = request.state.current_user
    is_base = getattr(request.state, "is_base_user", True)

    db_sess = db.query(ChatSession).filter(ChatSession.id == body.session_id).first()
    if db_sess:
        # ownership check — base users own their session; org users must share the same org
        if is_base and db_sess.user_id != user.id:
            raise HTTPException(status_code=403, detail="Access denied")
        if not is_base and db_sess.org_id != user.org_id:
            raise HTTPException(status_code=403, detail="Access denied")
    else:
        title = body.message[:30] + "..." if len(body.message) > 30 else body.message
        db_sess = ChatSession(id=body.session_id, user_id=user.id, org_id=user.org_id, title=title)
        db.add(db_sess)
        db.commit()

    blocked_msg = ChatMessage(session_id=body.session_id, role="blocked", content=body.message)
    db.add(blocked_msg)
    if body.model_explanation:
        model_msg = ChatMessage(session_id=body.session_id, role="model", content=body.model_explanation)
        db.add(model_msg)
    # use only type labels from client — do not trust raw values; real values were already
    # captured server-side in the StatLog created by /check
    detected = [t.get("type", "") for t in (body.blocked_types or []) if isinstance(t, dict)]
    db.add(StatLog(
        user_id=user.id,
        org_id=user.org_id,
        session_id=body.session_id,
        action="BLOCK",
        detected_types=json.dumps(detected),
        flagged_sequences=json.dumps([]),
        original_message=body.message,
    ))
    db.commit()
    return {"status": "saved"}

from app.models import BatchCheckRequest, BatchCheckResponse, CheckResult, BlockResult

@app.post("/api/v1/check_batch", response_model=BatchCheckResponse)
@limiter.limit("10/minute")
async def check_message_batch(request: Request, body: BatchCheckRequest, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    require_scope(request, "check")
    user = request.state.current_user
    tier_config, custom_labels = get_tier_config_helper(db, user)

    pipeline_results = await asyncio.gather(*[
        asyncio.to_thread(pipeline.run, msg, body.allowed_pii, [], tier_config, custom_labels)
        for msg in body.messages
    ])

    block_set = tier_config["block"] if tier_config else TIER_BLOCK
    results = []
    for msg, (processed_text, detections, action) in zip(body.messages, pipeline_results):
        db.add(StatLog(
            user_id=user.id,
            org_id=user.org_id,
            session_id=None,
            action=action,
            detected_types=json.dumps([d.type for d in detections]),
            flagged_sequences=json.dumps([msg[d.start:d.end] for d in detections]),
            original_message=msg if action in ["BLOCK", "REDACT"] else None
        ))
        redacted_types = [
            RedactedType(type=d.type, subtype=d.subtype, confidence=d.confidence)
            for d in detections
        ]
        if action == "BLOCK":
            blocked = [d.type for d in detections if d.type in block_set]
            results.append(BlockResult(
                action="BLOCK",
                warning=get_block_warning(blocked[0] if blocked else "code"),
                blocked_types=[rt for rt in redacted_types if rt.type in block_set]
            ))
        else:
            results.append(CheckResult(
                action=action,
                was_redacted=(action == "REDACT"),
                message=processed_text,
                redacted_types=redacted_types
            ))
    db.commit()
    return BatchCheckResponse(results=results)
@app.get("/api/v1/health")
def health():
    return {"status": "ok"}

def _resolve_org_user(user_id: str, current_user, db) -> "User | None":
    """Return a User in current_user's org by UUID string, employee_id, or email."""
    import uuid as _uuid
    try:
        uid = _uuid.UUID(user_id)
        return db.query(User).filter(User.id == uid, User.org_id == current_user.org_id).first()
    except ValueError:
        return db.query(User).filter(
            (User.employee_id == user_id) | (User.email == user_id),
            User.org_id == current_user.org_id
        ).first()


@app.get("/api/v1/admin/config/{user_id}", response_model=TierConfigResponse)
def get_user_config(request: Request, user_id: str, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    
    target_user = None
    if user_id in ["default_user", "me"]:
        if is_base_user:
            target_user = current_user
        else:
            if user_id == "default_user":
                from app.models_db import Organization
                org = db.query(Organization).filter(Organization.id == current_user.org_id).first()
                if not org or not org.default_tier_block:
                    return TierConfigResponse(
                        user_id=user_id,
                        tier_block=list(TIER_BLOCK),
                        tier_redact=list(TIER_REDACT),
                        tier_audit=list(TIER_AUDIT)
                    )
                return TierConfigResponse(
                    user_id=user_id,
                    tier_block=org.default_tier_block,
                    tier_redact=org.default_tier_redact,
                    tier_audit=org.default_tier_audit
                )
            else:
                target_user = current_user
    else:
        if is_base_user:
            target_user = current_user
        else:
            target_user = _resolve_org_user(user_id, current_user, db)
            if not target_user:
                raise HTTPException(status_code=404, detail="User not found")

    if not target_user or not target_user.tier_block:
        return TierConfigResponse(
            user_id=user_id,
            tier_block=list(TIER_BLOCK),
            tier_redact=list(TIER_REDACT),
            tier_audit=list(TIER_AUDIT)
        )
    return TierConfigResponse(
        user_id=user_id,
        tier_block=target_user.tier_block,
        tier_redact=target_user.tier_redact,
        tier_audit=target_user.tier_audit
    )

@app.post("/api/v1/admin/config/{user_id}")
def update_user_config(request: Request, user_id: str, body: TierConfigUpdate, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    
    if user_id in ["default_user", "me"]:
        if is_base_user:
            current_user.tier_block = body.tier_block
            current_user.tier_redact = body.tier_redact
            current_user.tier_audit = body.tier_audit
            db.add(current_user)
            db.commit()
            return {"status": "updated"}
        else:
            if user_id == "default_user":
                from app.models_db import Organization
                org = db.query(Organization).filter(Organization.id == current_user.org_id).first()
                if org:
                    org.default_tier_block = body.tier_block
                    org.default_tier_redact = body.tier_redact
                    org.default_tier_audit = body.tier_audit
                    db.add(org)
                    db.commit()
                    return {"status": "updated"}
                raise HTTPException(status_code=404, detail="Organization not found")
            else:
                current_user.tier_block = body.tier_block
                current_user.tier_redact = body.tier_redact
                current_user.tier_audit = body.tier_audit
                db.add(current_user)
                db.commit()
                return {"status": "updated"}
    else:
        if is_base_user:
            target_user = current_user
        else:
            target_user = _resolve_org_user(user_id, current_user, db)
            if not target_user:
                raise HTTPException(status_code=404, detail="User not found")

        target_user.tier_block = body.tier_block
        target_user.tier_redact = body.tier_redact
        target_user.tier_audit = body.tier_audit
        db.add(target_user)
        db.commit()
        return {"status": "updated"}

from typing import Optional
from datetime import datetime

def fetch_stats(db, current_user, is_base_user: bool, user_id=None, start_time=None, end_time=None):
    if user_id == "me":
        user_id = str(current_user.id)

    query = db.query(StatLog)
    if is_base_user:
        query = query.filter(StatLog.user_id == current_user.id)
    else:
        query = query.filter(StatLog.org_id == current_user.org_id)
        
    if user_id and user_id not in ["default_user", "me"]:
        if is_base_user:
            query = query.filter(StatLog.user_id == current_user.id)
        else:
            target_user = _resolve_org_user(user_id, current_user, db)
            if not target_user:
                return StatsResponse(total_requests=0, actions=[], detected_types=[], top_sequences=[])
            query = query.filter(StatLog.user_id == target_user.id)
            
    if start_time:
        query = query.filter(StatLog.created_at >= start_time)
    if end_time:
        query = query.filter(StatLog.created_at <= end_time)
        
    total = query.with_entities(func.count(StatLog.id)).scalar() or 0
    actions = query.with_entities(StatLog.action, func.count(StatLog.id)).group_by(StatLog.action).all()

    total_tokens = query.with_entities(func.coalesce(func.sum(StatLog.tokens), 0)).scalar() or 0  # NULL = interrupted stream, flagged not estimated
    untracked = query.filter(
        StatLog.tokens.is_(None),
        StatLog.action != "BLOCK",
        StatLog.api_key_id.is_(None),
    ).with_entities(func.count(StatLog.id)).scalar() or 0

    type_counts = Counter()
    seq_counts = Counter()
    for log in query.with_entities(StatLog.detected_types, StatLog.flagged_sequences).all():
        types_val = log[0]
        if isinstance(types_val, str):
            try:
                types_val = json.loads(types_val)
            except:
                types_val = []
        if isinstance(types_val, list):
            type_counts.update(types_val)
            
        seq_val = log[1]
        if seq_val:
            if isinstance(seq_val, str):
                try:
                    seq_val = json.loads(seq_val)
                except:
                    seq_val = []
            if isinstance(seq_val, list):
                seq_counts.update(seq_val)
    return StatsResponse(
        total_requests=total,
        actions=[StatCount(name=a[0], count=a[1]) for a in actions],
        detected_types=[StatCount(name=k, count=v) for k, v in type_counts.items()],
        top_sequences=[StatCount(name=k, count=v) for k, v in seq_counts.most_common(10)],
        total_tokens=int(total_tokens),
        tokens_incomplete=untracked > 0,
    )

@app.get("/api/v1/admin/users")
def list_org_users(request: Request, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    if is_base_user:
        return []
    users = db.query(User).filter(
        User.org_id == current_user.org_id,
        User.is_active == True,
        # Exclude Clerk user IDs stored as email (they look like "user_xxx", not real emails)
        User.email.contains('@') | User.employee_id.isnot(None)
    ).order_by(User.created_at).all()
    result = []
    for u in users:
        # employee_id holds the real email when Clerk JWT provides it; fall back to email column
        display_email = u.employee_id if (u.employee_id and '@' in u.employee_id) else (u.email if '@' in (u.email or '') else None)
        if not display_email:
            continue
        result.append({
            "id": str(u.id),
            "email": display_email,
            "employee_id": u.employee_id or "",
            "role": u.role,
            "is_blocked": u.is_blocked,
        })
    return result

@app.get("/api/v1/admin/stats", response_model=StatsResponse)
def get_global_stats(request: Request, start_time: Optional[datetime] = None, end_time: Optional[datetime] = None, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    return fetch_stats(db, current_user, is_base_user, None, start_time, end_time)

@app.get("/api/v1/admin/stats/{user_id}", response_model=StatsResponse)
def get_user_stats(request: Request, user_id: str, start_time: Optional[datetime] = None, end_time: Optional[datetime] = None, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    return fetch_stats(db, current_user, is_base_user, user_id, start_time, end_time)

@app.get("/api/v1/admin/logs/export")
def export_audit_log(request: Request, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    """Download the audit trail as CSV, scoped to the caller's org (or self for solo users)."""
    import csv, io
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)

    q = db.query(StatLog)
    if is_base_user:
        q = q.filter(StatLog.user_id == current_user.id)
    else:
        q = q.filter(StatLog.org_id == current_user.org_id)
    logs = q.order_by(StatLog.created_at.desc()).limit(10000).all()

    def _list(v):
        if isinstance(v, str):
            try:
                v = json.loads(v)
            except Exception:
                return ""
        return ", ".join(v) if isinstance(v, list) else ""

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["timestamp", "user_id", "action", "detected_types", "flagged_sequences", "via_api_key"])
    for log in logs:
        w.writerow([
            log.created_at.isoformat() if log.created_at else "",
            str(log.user_id),
            log.action,
            _list(log.detected_types),
            _list(log.flagged_sequences),
            "yes" if log.api_key_id else "no",
        ])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=adopshun_audit_log.csv"},
    )


@app.get("/api/v1/admin/logs/{user_id}")
def get_user_logs(request: Request, user_id: str, limit: int = 50, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    
    if is_base_user:
        logs = db.query(StatLog).filter(StatLog.user_id == current_user.id).order_by(StatLog.created_at.desc()).limit(limit).all()
    else:
        if user_id == "me":
            logs = db.query(StatLog).filter(StatLog.user_id == current_user.id).order_by(StatLog.created_at.desc()).limit(limit).all()
        elif user_id == "default_user":
            logs = db.query(StatLog).filter(StatLog.org_id == current_user.org_id).order_by(StatLog.created_at.desc()).limit(limit).all()
        else:
            target_user = _resolve_org_user(user_id, current_user, db)
            if not target_user:
                raise HTTPException(status_code=404, detail="User not found")
            logs = db.query(StatLog).filter(StatLog.user_id == target_user.id).order_by(StatLog.created_at.desc()).limit(limit).all()

    result_logs = []
    for log in logs:
        types_val = log.detected_types
        if isinstance(types_val, str):
            try:
                types_val = json.loads(types_val)
            except:
                types_val = []
                
        seq_val = log.flagged_sequences
        if seq_val:
            if isinstance(seq_val, str):
                try:
                    seq_val = json.loads(seq_val)
                except:
                    seq_val = []
            else:
                seq_val = []
        else:
            seq_val = []
            
        result_logs.append({
            "id": log.id,
            "action": log.action,
            "detected_types": types_val,
            "flagged_sequences": seq_val,
            "original_message": log.original_message,
            "created_at": log.created_at.isoformat()
        })
    return result_logs


@app.post("/api/v1/admin/retention/purge")
def trigger_retention_purge(request: Request, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    """Manually enforce retention now. Org admins purge only their own org."""
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    if is_base_user:
        return {"status": "skipped", "reason": "retention is an organisation policy"}
    deleted = purge_old_data(db, org_id=current_user.org_id)
    return {"status": "ok", "deleted_stat_logs": deleted}


@app.get("/api/v1/patterns")
def list_patterns(_: bool = Depends(verify_credentials)):
    return {
        "stages": [
            "regex_recognisers",
            "structural_recognisers",
            "entropy_analyser",
            "luhn_validator",
        ],
        "regex_patterns": [p.name for p in regex_stage._PATTERNS],
    }

from app.models import CustomLabelCreate, CustomLabelResponse

BUILT_IN_LABEL_METADATA = {
    "api_key":          {"description": "API keys, tokens, and credentials used to authenticate with third-party services.", "examples": ["sk-abc123...", "ghp_xxxx", "eyJhbGci... (JWT)", "AKIAIOSFODNN7 (AWS)"], "detection_methods": ["Regex patterns", "Entropy analysis"]},
    "password":         {"description": "Passwords and secrets found in key=value or key: value patterns.", "examples": ["password=secret123", "pwd: mypass", "secret: abc"], "detection_methods": ["Regex (key-value patterns)"]},
    "credit_card":      {"description": "Credit and debit card numbers (Visa, Mastercard, Amex, Discover).", "examples": ["4111-1111-1111-1111", "5500 0000 0000 0004"], "detection_methods": ["Regex", "Luhn checksum"]},
    "card number":      {"description": "Payment card numbers in various formats.", "examples": ["4111 1111 1111 1111", "3714 496353 98431"], "detection_methods": ["Regex", "Luhn checksum"]},
    "credit card":      {"description": "Credit card numbers in standard formats.", "examples": ["4111-1111-1111-1111", "5500-0000-0000-0004"], "detection_methods": ["Regex", "Luhn checksum"]},
    "CVV":              {"description": "Card verification values — 3 or 4 digit security codes on payment cards.", "examples": ["CVV: 123", "CVC: 4321"], "detection_methods": ["Regex"]},
    "IBAN":             {"description": "International Bank Account Numbers used for wire transfers.", "examples": ["GB29 NWBK 6016 1331 9268 19", "DE89 3704 0044 0532 0130 00"], "detection_methods": ["Regex"]},
    "IBAN code":        {"description": "IBAN codes in condensed format.", "examples": ["GB29NWBK60161331926819"], "detection_methods": ["Regex"]},
    "US bank number":   {"description": "US bank routing numbers (ABA/transit numbers).", "examples": ["routing: 021000021", "ABA: 111000025"], "detection_methods": ["Regex"]},
    "API keys":         {"description": "Generic API keys from platforms like Stripe, Slack, Google.", "examples": ["sk_live_xxx (Stripe)", "xoxb-xxx (Slack)", "AIza-xxx (Google)"], "detection_methods": ["Regex"]},
    "code":             {"description": "Code snippets and scripts that may leak business logic or enable injection.", "examples": ["function foo() {...}", "SELECT * FROM users", "import os; os.system(...)"], "detection_methods": ["Code density analysis", "Regex"]},
    "private_key":      {"description": "Private cryptographic keys (RSA, EC, SSH, PGP).", "examples": ["-----BEGIN RSA PRIVATE KEY-----", "-----BEGIN EC PRIVATE KEY-----"], "detection_methods": ["Regex"]},
    "crypto wallet":    {"description": "Cryptocurrency wallet addresses (Bitcoin, Ethereum).", "examples": ["1A1zP1eP5QGefi2DMPTfTL5SLmv7...", "0x742d35Cc6634C0532925a3b..."], "detection_methods": ["Regex"]},
    "email":            {"description": "Email addresses in standard RFC format.", "examples": ["user@example.com", "john.doe+tag@company.org"], "detection_methods": ["Regex"]},
    "ssn":              {"description": "US Social Security Numbers.", "examples": ["123-45-6789", "987 65 4321"], "detection_methods": ["Regex"]},
    "US SSN":           {"description": "US Social Security Numbers in standard dashed format.", "examples": ["123-45-6789"], "detection_methods": ["Regex"]},
    "passport number":  {"description": "Passport numbers and travel document identifiers.", "examples": ["A12345678", "P<USASMITH"], "detection_methods": ["Regex", "AI (Gemini)"]},
    "US passport":      {"description": "US passport numbers (letter + 8 digits).", "examples": ["A12345678", "B98765432"], "detection_methods": ["AI (Gemini)"]},
    "driver's license": {"description": "Driver's license numbers from various states and countries.", "examples": ["DL: A1234567", "License: 12345678"], "detection_methods": ["AI (Gemini)"]},
    "US driver license":{"description": "US state driver's license numbers.", "examples": ["CA: A1234567", "TX: 12345678"], "detection_methods": ["AI (Gemini)"]},
    "US ITIN":          {"description": "US Individual Taxpayer Identification Numbers (9XX-XX-XXXX format).", "examples": ["900-70-0000", "976-45-6789"], "detection_methods": ["AI (Gemini)"]},
    "UK NHS number":    {"description": "UK National Health Service patient identifiers.", "examples": ["NHS: 943 476 5919", "national health: 123 456 7890"], "detection_methods": ["Regex"]},
    "tax ID":           {"description": "Tax identification numbers (Aadhaar, PAN, Voter ID).", "examples": ["1234 5678 9012 (Aadhaar)", "ABCDE1234F (PAN)", "ABC1234567 (Voter ID)"], "detection_methods": ["Regex"]},
    "full_name":        {"description": "Full names of individuals.", "examples": ["John Smith", "Jane Doe"], "detection_methods": ["AI (Gemini)"]},
    "person":           {"description": "Person names and personal identifiers.", "examples": ["Mr. John Smith", "Dr. Jane Doe"], "detection_methods": ["AI (Gemini)"]},
    "location":         {"description": "Geographic locations, place names, and landmarks.", "examples": ["New York", "Eiffel Tower", "San Francisco, CA"], "detection_methods": ["AI (Gemini)"]},
    "organization":     {"description": "Organization and company names.", "examples": ["Acme Corp", "Microsoft", "Google LLC"], "detection_methods": ["AI (Gemini)"]},
    "phone number":     {"description": "Phone numbers in US and Indian formats.", "examples": ["+1 (555) 123-4567", "+91 98765 43210"], "detection_methods": ["Regex"]},
    "physical address": {"description": "Physical mailing and street addresses.", "examples": ["123 Main Street, Apt 4B, New York, NY 10001"], "detection_methods": ["AI (Gemini)"]},
    "IP address":       {"description": "IPv4 addresses in decimal or hex format.", "examples": ["192.168.1.1", "10.0.0.1", "0xC0A80101"], "detection_methods": ["Regex"]},
}

@app.get("/api/v1/admin/labels/all")
def get_all_labels(request: Request, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)

    if is_base_user:
        custom_labels = db.query(CustomLabel).filter(CustomLabel.user_id == current_user.id).all()
    else:
        custom_labels = db.query(CustomLabel).filter(CustomLabel.org_id == current_user.org_id).all()

    custom_map = {cl.name: cl for cl in custom_labels}
    built_in_names = TIER_BLOCK | TIER_REDACT | TIER_AUDIT
    all_labels = []

    def _parse_words(raw) -> list:
        if not raw:
            return []
        try:
            result = json.loads(raw) if isinstance(raw, str) else raw
            return result if isinstance(result, list) else []
        except Exception:
            return []

    for tier_name, tier_set in [("tier_block", TIER_BLOCK), ("tier_redact", TIER_REDACT), ("tier_audit", TIER_AUDIT)]:
        for name in tier_set:
            meta = BUILT_IN_LABEL_METADATA.get(name, {})
            cl = custom_map.get(name)
            words = _parse_words(cl.dictionary_words) if cl else []
            all_labels.append({
                "id": cl.id if cl else None,
                "name": name,
                "description": (cl.description if cl and cl.description else None) or meta.get("description", ""),
                "tier": tier_name,
                "dictionary_words": words,
                "is_builtin": True,
                "detection_methods": meta.get("detection_methods", []),
                "examples": meta.get("examples", []),
            })

    for cl in custom_labels:
        if cl.name not in built_in_names:
            words = _parse_words(cl.dictionary_words)
            all_labels.append({
                "id": cl.id,
                "name": cl.name,
                "description": cl.description or "",
                "tier": cl.tier,
                "dictionary_words": words,
                "is_builtin": False,
                "detection_methods": ["Dictionary", "AI (Gemini)"],
                "examples": words[:3],
            })

    return all_labels

@app.post("/api/v1/admin/custom_labels", response_model=CustomLabelResponse)
def create_custom_label(request: Request, label: CustomLabelCreate, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    user = request.state.current_user if hasattr(request.state, 'current_user') else None
    is_base_user = getattr(request.state, "is_base_user", True)
    
    if is_base_user:
        existing = db.query(CustomLabel).filter(CustomLabel.name == label.name, CustomLabel.user_id == user.id).first()
    else:
        existing = db.query(CustomLabel).filter(CustomLabel.name == label.name, CustomLabel.org_id == user.org_id).first()
        
    if existing:
        raise HTTPException(status_code=400, detail="Label already exists")
    
    dump_data = label.model_dump()
    dump_data["dictionary_words"] = json.dumps(dump_data.get("dictionary_words", []))
    
    if _gemini_client:
        try:
            client = _gemini_client
            words_context = ""
            if label.dictionary_words:
                words_context = f" Examples to match: {', '.join(label.dictionary_words)}."
            prompt = f"Generate ONLY a python regex string to match the PII entity '{label.name}'. Description: '{label.description}'.{words_context} Do not include markdown blocks, slashes, or explanations. Just the raw regex pattern. Make it strict to avoid false positives."
            response = client.models.generate_content(
                model='gemini-3.5-flash',
                contents=prompt,
            )
            dump_data["regex_pattern"] = response.text.strip().strip('`').strip()
        except Exception as e:
            print("Gemini Regex Generation Failed:", e)

    if dump_data.get("regex_pattern"):
        _validate_regex_safe(dump_data["regex_pattern"])
            
    db_label = CustomLabel(
        name=dump_data.get("name"),
        description=dump_data.get("description"),
        tier=dump_data.get("tier"),
        regex_pattern=dump_data.get("regex_pattern"),
        scope="user" if is_base_user else "org",
        user_id=user.id if is_base_user else None,
        org_id=None if is_base_user else user.org_id,
        dictionary_words=dump_data.get("dictionary_words", "[]")
    )
    db.add(db_label)
    db.commit()
    db.refresh(db_label)
    
    try:
        words = json.loads(db_label.dictionary_words)
    except:
        words = []
        
    return {
        "id": db_label.id,
        "name": db_label.name,
        "description": db_label.description,
        "tier": db_label.tier,
        "regex_pattern": db_label.regex_pattern,
        "dictionary_words": words,
        "created_at": db_label.created_at
    }

@app.get("/api/v1/admin/custom_labels", response_model=list[CustomLabelResponse])
def get_custom_labels(request: Request, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    
    if is_base_user:
        labels = db.query(CustomLabel).filter(CustomLabel.user_id == current_user.id).all()
    else:
        labels = db.query(CustomLabel).filter(CustomLabel.org_id == current_user.org_id).all()
        
    results = []
    for lbl in labels:
        try:
            words = json.loads(lbl.dictionary_words)
        except:
            words = []
        results.append({
            "id": lbl.id,
            "name": lbl.name,
            "description": lbl.description,
            "tier": lbl.tier,
            "regex_pattern": lbl.regex_pattern,
            "dictionary_words": words,
            "created_at": lbl.created_at
        })
    return results

@app.delete("/api/v1/admin/custom_labels/{label_id}")
def delete_custom_label(request: Request, label_id: int, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    
    db_label = db.query(CustomLabel).filter(CustomLabel.id == label_id).first()
    if not db_label:
        raise HTTPException(status_code=404, detail="Label not found")
        
    if is_base_user and db_label.user_id != user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    if not is_base_user and db_label.org_id != user.org_id:
        raise HTTPException(status_code=403, detail="Access denied")
        
    db.delete(db_label)
    db.commit()
    return {"status": "deleted"}

from pydantic import BaseModel
class CustomLabelUpdate(BaseModel):
    dictionary_words: list[str]

@app.put("/api/v1/admin/custom_labels/{label_id}", response_model=CustomLabelResponse)
def update_custom_label(request: Request, label_id: int, update: CustomLabelUpdate, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    
    db_label = db.query(CustomLabel).filter(CustomLabel.id == label_id).first()
    if not db_label:
        raise HTTPException(status_code=404, detail="Label not found")
        
    if is_base_user and db_label.user_id != user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    if not is_base_user and db_label.org_id != user.org_id:
        raise HTTPException(status_code=403, detail="Access denied")
        
    db_label.dictionary_words = json.dumps(update.dictionary_words)
    db.commit()
    db.refresh(db_label)
    
    try:
        words = json.loads(db_label.dictionary_words)
    except:
        words = []
        
    return {
        "id": db_label.id,
        "name": db_label.name,
        "description": db_label.description,
        "tier": db_label.tier,
        "regex_pattern": db_label.regex_pattern,
        "dictionary_words": words,
        "created_at": db_label.created_at
    }

from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse
import csv
import io

@app.get("/api/v1/admin/custom_labels/xlsx")
def export_custom_labels_xlsx(request: Request, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    
    if is_base_user:
        labels = db.query(CustomLabel).filter(CustomLabel.user_id == current_user.id).all()
    else:
        labels = db.query(CustomLabel).filter(CustomLabel.org_id == current_user.org_id).all()
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entity Labels"
    
    headers = ["Name", "Description", "Tier", "Dictionary Words"]
    ws.append(headers)

    for lbl in labels:
        try:
            words = json.loads(lbl.dictionary_words) if lbl.dictionary_words else []
        except Exception:
            words = []
        ws.append([lbl.name, lbl.description, lbl.tier, ", ".join(words) if words else ""])
        
    dv = DataValidation(type="list", formula1='"tier_block,tier_redact,tier_audit"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add('C2:C1000')
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=entity_labels.xlsx"}
    )

@app.get("/api/v1/admin/dictionary/xlsx")
def export_dictionary_xlsx(request: Request, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    
    if is_base_user:
        labels = db.query(CustomLabel).filter(CustomLabel.user_id == current_user.id).all()
    else:
        labels = db.query(CustomLabel).filter(CustomLabel.org_id == current_user.org_id).all()
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dictionary Labels"
    
    headers = ["Name", "Description", "Tier", "Dictionary Words"]
    ws.append(headers)
    
    for lbl in labels:
        try:
            words = json.loads(lbl.dictionary_words)
        except:
            words = []
        ws.append([lbl.name, lbl.description, lbl.tier, ", ".join(words)])
        
    dv = DataValidation(type="list", formula1='"tier_block,tier_redact,tier_audit"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add('C2:C1000')
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dictionary_labels.xlsx"}
    )

@app.post("/api/v1/admin/custom_labels/import/preview")
async def import_custom_labels_xlsx_preview(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    import os
    from pydantic import BaseModel
    
    class TierResponse(BaseModel):
        tier: str
        
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    
    rows = list(ws.rows)
    if not rows:
        raise HTTPException(status_code=400, detail="Empty Excel file.")
        
    header = [cell.value for cell in rows[0]]
    if "Name" not in header:
        raise HTTPException(status_code=400, detail="Invalid Excel format. Missing 'Name' header.")
        
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    
    client = _gemini_client
    
    results = []
    
    for row in rows[1:]:
        row_values = [cell.value for cell in row]
        if not row_values or not row_values[0]:
            continue
            
        name = str(row_values[0]).strip()
        description = str(row_values[1]).strip() if len(row_values) > 1 and row_values[1] else ""
        tier = str(row_values[2]).strip() if len(row_values) > 2 and row_values[2] else ""
        words_str = str(row_values[3]).strip() if len(row_values) > 3 and row_values[3] else ""
        
        words = [w.strip() for w in words_str.split(",")] if words_str else []
        
        if is_base_user:
            existing = db.query(CustomLabel).filter(CustomLabel.name == name, CustomLabel.user_id == current_user.id).first()
        else:
            existing = db.query(CustomLabel).filter(CustomLabel.name == name, CustomLabel.org_id == current_user.org_id).first()
        is_new = existing is None
        
        if not tier or tier not in ["tier_block", "tier_redact", "tier_audit"]:
            if client:
                try:
                    prompt = f"Assign a tier for the PII entity '{name}'. Description: '{description}'. You MUST choose one of: tier_block, tier_redact, tier_audit."
                    response = client.models.generate_content(
                        model='gemini-3.5-flash',
                        contents=prompt,
                        config=types.GenerateContentConfig(
                            response_mime_type="application/json",
                            response_schema=TierResponse,
                            temperature=0.0
                        )
                    )
                    tier_data = json.loads(response.text)
                    tier = tier_data.get("tier", "tier_audit")
                    if tier not in ["tier_block", "tier_redact", "tier_audit"]:
                        tier = "tier_audit"
                except Exception as e:
                    print("Gemini Tier Prediction Failed:", e)
                    tier = "tier_audit"
            else:
                tier = "tier_audit"
                
        results.append({
            "name": name,
            "description": description,
            "tier": tier,
            "dictionary_words": words,
            "is_new": is_new
        })
        
    return {"preview": results}

class ImportConfirmPayload(BaseModel):
    items: list[dict]

@app.post("/api/v1/admin/custom_labels/import/confirm")
def import_custom_labels_xlsx_confirm(request: Request, payload: ImportConfirmPayload, db: Session = Depends(get_db), _: bool = Depends(verify_credentials)):
    import os
    
    current_user = request.state.current_user
    is_base_user = getattr(request.state, "is_base_user", True)
    
    client = _gemini_client
    
    imported_count = 0
    for item in payload.items:
        name = item.get("name")
        description = item.get("description", "")
        tier = item.get("tier", "tier_audit")
        words = item.get("dictionary_words", [])
        
        if is_base_user:
            existing = db.query(CustomLabel).filter(CustomLabel.name == name, CustomLabel.user_id == current_user.id).first()
        else:
            existing = db.query(CustomLabel).filter(CustomLabel.name == name, CustomLabel.org_id == current_user.org_id).first()
            
        if existing:
            existing.description = description
            existing.tier = tier
            if words:  # blank D column on import must not wipe existing dictionary words
                existing.dictionary_words = json.dumps(words)
        else:
            regex_pattern = ""
            if not words and client:
                try:
                    prompt = f"Generate ONLY a python regex string to match the PII entity '{name}'. Description: '{description}'. Do not include markdown blocks, slashes, or explanations. Just the raw regex pattern."
                    response = client.models.generate_content(
                        model='gemini-3.5-flash',
                        contents=prompt,
                    )
                    regex_pattern = response.text.strip()
                except Exception as e:
                    print("Gemini Regex Generation Failed:", e)
                    
            db_label = CustomLabel(
                name=name,
                description=description,
                tier=tier,
                scope="user" if is_base_user else "org",
                user_id=current_user.id if is_base_user else None,
                org_id=None if is_base_user else current_user.org_id,
                dictionary_words=json.dumps(words),
                regex_pattern=regex_pattern
            )
            db.add(db_label)
        imported_count += 1
        
    db.commit()
    return {"status": "success", "imported": imported_count}

from fastapi import Form
from app.pipeline.document_stage import extract_text
import tempfile
import os

_MAGIC_BYTES: dict[str, bytes] = {
    ".pdf":  b"%PDF",
    ".docx": b"PK\x03\x04",
    ".xlsx": b"PK\x03\x04",
    ".pptx": b"PK\x03\x04",
    ".jpg":  b"\xff\xd8",
    ".jpeg": b"\xff\xd8",
    ".png":  b"\x89PNG",
    ".bmp":  b"BM",
    ".gif":  b"GIF8",
    ".tiff": b"II*\x00",
    ".tif":  b"II*\x00",
    ".webp": b"RIFF",
}
_UPLOAD_FLAG: dict[str, str] = {
    ".pdf":  "can_upload_pdf",
    ".jpg":  "can_upload_image", ".jpeg": "can_upload_image",
    ".png":  "can_upload_image", ".bmp":  "can_upload_image",
    ".tiff": "can_upload_image", ".tif":  "can_upload_image",
    ".webp": "can_upload_image",
    ".csv":  "can_upload_csv",
    ".docx": "can_upload_docx",
}

def _validate_upload(content: bytes, suffix: str, user) -> None:
    magic = _MAGIC_BYTES.get(suffix)
    if magic and not content[:len(magic)] == magic:
        raise HTTPException(status_code=415, detail="File content does not match its extension.")
    flag = _UPLOAD_FLAG.get(suffix)
    if flag and not getattr(user, flag, True):
        raise HTTPException(status_code=403, detail="Your account is not permitted to upload this file type.")

@app.post("/api/v1/document/upload", response_model=Union[CheckResult, BlockResult])
@limiter.limit("10/minute")
async def upload_document(
    request: Request, 
    file: UploadFile = File(...), 
    session_id: str = Form("default_session"),
    db: Session = Depends(get_db),
    _: bool = Depends(verify_credentials)
):
    require_scope(request, "check")
    user = request.state.current_user
    tier_config, custom_labels = get_tier_config_helper(db, user)

    max_bytes = int(os.getenv("MAX_UPLOAD_MB", "50")) * 1024 * 1024  # must match nginx client_max_body_size
    content_length = request.headers.get("content-length")
    if content_length and int(content_length) > max_bytes:
        raise HTTPException(status_code=413, detail=f"File too large. Maximum size is {max_bytes // (1024 * 1024)}MB.")

    _ALLOWED_EXTS = {'.pdf', '.docx', '.xlsx', '.pptx', '.txt', '.csv', '.md',
                     '.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.webp'}
    safe_filename = os.path.basename(file.filename or "upload")
    suffix = os.path.splitext(safe_filename)[1].lower() if safe_filename else ".tmp"
    if suffix not in _ALLOWED_EXTS:
        raise HTTPException(status_code=415, detail=f"Unsupported file type '{suffix}'. Allowed: pdf, docx, xlsx, pptx, txt, csv, images.")
    content = await file.read()
    if len(content) > max_bytes:
        raise HTTPException(status_code=413, detail=f"File too large. Maximum size is {max_bytes // (1024 * 1024)}MB.")
    _validate_upload(content, suffix, user)
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp_path = tmp.name
        tmp.write(content)
        
    try:
        extracted_text = await asyncio.to_thread(extract_text, tmp_path)  # OCR is CPU-heavy; keep off the event loop
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    if not extracted_text.strip():
        ext = os.path.splitext(file.filename or "")[1].lower()
        if ext in {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.webp'}:
            raise HTTPException(
                status_code=422,
                detail="No readable text found in the image. It may be blurry, low-resolution, or contain no text. Try a clearer image.",
            )
        raise HTTPException(status_code=400, detail="Could not extract text from document, or document was empty.")
        
    processed_text, detections, action = await asyncio.to_thread(pipeline.run, extracted_text, [], [], tier_config, custom_labels)
    
    import uuid
    try:
        real_sess_id = uuid.UUID(session_id)
    except (ValueError, AttributeError):
        real_sess_id = None

    detected_types_list = [d.type for d in detections]
    flagged_sequences_list = [extracted_text[d.start:d.end] for d in detections]

    if real_sess_id:
        db_sess = db.query(ChatSession).filter(ChatSession.id == real_sess_id).first()
        if not db_sess:
            db_sess = ChatSession(
                id=real_sess_id,
                user_id=user.id,
                org_id=user.org_id,
                title=f"Doc: {safe_filename[:40]}"
            )
            db.add(db_sess)

    stat_log = StatLog(
        user_id=user.id,
        org_id=user.org_id,
        session_id=real_sess_id,
        action=action,
        detected_types=json.dumps(detected_types_list),
        flagged_sequences=json.dumps(flagged_sequences_list)
    )
    db.add(stat_log)
    db.commit()

    redacted_types = [
        RedactedType(type=d.type, subtype=d.subtype, confidence=d.confidence, value=extracted_text[d.start:d.end])
        for d in detections
    ]
    
    if action == "BLOCK":
        block_set = tier_config["block"] if tier_config else TIER_BLOCK
        blocked_types = [d.type for d in detections if d.type in block_set]
        primary_type = blocked_types[0] if blocked_types else "code"

        blocked_redacted_types = [rt for rt in redacted_types if rt.type in block_set]

        return BlockResult(
            action="BLOCK",
            warning=get_block_warning(primary_type),
            blocked_types=blocked_redacted_types
        )

    block_set = tier_config["block"] if tier_config else TIER_BLOCK
    tokenized, vault = tokenization.build_tokens(extracted_text, detections, block_set)

    return CheckResult(
        action=action,
        was_redacted=(action == "REDACT"),
        message=processed_text,
        redacted_types=redacted_types,
        tokenized=tokenized,
        vault=vault,
    )


@app.post("/api/v1/admin/invite/bulk")
async def bulk_invite(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(verify_credentials),
):
    import csv, io
    from clerk_backend_api import Clerk
    from clerk_backend_api.models import CreateOrganizationInvitationBulkRequestBody

    clerk_org_id = getattr(request.state, "clerk_org_id", None)
    clerk_user_id = getattr(request.state, "clerk_user_id", None)
    if not clerk_org_id:
        raise HTTPException(status_code=400, detail="No active organization in session")

    content = (await file.read()).decode("utf-8-sig")
    reader = csv.reader(io.StringIO(content))
    emails = []
    for i, row in enumerate(reader):
        if not row:
            continue
        val = row[0].strip()
        if i == 0 and ("email" in val.lower() or "@" not in val):
            continue  # skip header row
        if "@" in val:
            emails.append(val.lower())

    if not emails:
        raise HTTPException(status_code=400, detail="No valid email addresses found in CSV")

    clerk = Clerk(bearer_auth=os.getenv("CLERK_SECRET_KEY"))
    sent, failed = [], []

    for i in range(0, len(emails), 10):
        batch = emails[i:i + 10]
        try:
            clerk.organization_invitations.bulk_create(
                organization_id=clerk_org_id,
                request_body=[
                    CreateOrganizationInvitationBulkRequestBody(
                        email_address=email,
                        role="org:member",
                        inviter_user_id=clerk_user_id,
                        redirect_url="https://chat.adopshun.com",
                    )
                    for email in batch
                ],
            )
            sent.extend(batch)
        except Exception as e:
            logger.warning("Bulk invite batch failed: %s", e)
            failed.extend(batch)

    return {"sent": sent, "failed": failed, "total": len(emails)}
